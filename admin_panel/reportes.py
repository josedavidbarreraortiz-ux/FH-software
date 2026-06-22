import io
import csv
from datetime import datetime, date
from decimal import Decimal, InvalidOperation

from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Q, F, Sum, Count

from accounts.models import (
    User, Producto, Categoria, Inventario, Venta, VentaDetalle,
    MovimientoInventario, MetodoPago
)
from admin_panel.views import admin_required, _get_marcas
import bcrypt

# ─── ReportLab imports ────────────────────────────────────────────
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch, mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT


# ══════════════════════════════════════════════════════════════════
#  ESTILOS PDF REUTILIZABLES
# ══════════════════════════════════════════════════════════════════

def _get_pdf_styles():
    styles = getSampleStyleSheet()
    return {
        'title': ParagraphStyle('T', parent=styles['Title'], fontSize=20,
                                textColor=colors.HexColor('#0d47a1'), spaceAfter=6),
        'subtitle': ParagraphStyle('S', parent=styles['Normal'], fontSize=10,
                                   textColor=colors.HexColor('#666666'), alignment=TA_CENTER, spaceAfter=4),
        'cell': ParagraphStyle('C', parent=styles['Normal'], fontSize=8, leading=10),
        'cell_center': ParagraphStyle('CC', parent=styles['Normal'], fontSize=8, leading=10, alignment=TA_CENTER),
        'cell_right': ParagraphStyle('CR', parent=styles['Normal'], fontSize=8, leading=10, alignment=TA_RIGHT),
        'filtro': ParagraphStyle('F', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#888888')),
    }


def _build_pdf(elements, orientation='landscape'):
    buffer = io.BytesIO()
    pagesize = landscape(letter) if orientation == 'landscape' else letter
    doc = SimpleDocTemplate(buffer, pagesize=pagesize,
                            topMargin=20*mm, bottomMargin=15*mm, leftMargin=15*mm, rightMargin=15*mm)
    doc.build(elements)
    buffer.seek(0)
    return buffer


def _table_style():
    return TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d47a1')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f7ff')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ])


def _pdf_header(elements, titulo, subtitulo, total_text, filtros=None):
    s = _get_pdf_styles()
    elements.append(Paragraph('FH TechStore', s['title']))
    elements.append(Paragraph(titulo, s['subtitle']))
    now = datetime.now().strftime('%d/%m/%Y %H:%M')
    elements.append(Paragraph(f'Generado: {now} — {total_text}', s['subtitle']))
    elements.append(Spacer(1, 10))
    if filtros:
        elements.append(Paragraph(f'Filtros: {" | ".join(filtros)}', s['filtro']))
        elements.append(Spacer(1, 8))


# ══════════════════════════════════════════════════════════════════
#  PÁGINA PRINCIPAL DE REPORTES
# ══════════════════════════════════════════════════════════════════

@admin_required
def reportes_index(request):
    return render(request, 'admin_panel/reportes_index.html')


# ══════════════════════════════════════════════════════════════════
#  REPORTE DE PRODUCTOS
# ══════════════════════════════════════════════════════════════════

@admin_required
def reporte_productos_view(request):
    categorias = Categoria.objects.filter(estado=1).order_by('categoria_nombre')
    marcas = _get_marcas()
    return render(request, 'admin_panel/reporte_productos.html', {
        'categorias': categorias,
        'marcas': marcas,
    })


@admin_required
def reporte_productos_pdf(request):
    producto_id = request.GET.get('id', '').strip()
    nombre = request.GET.get('nombre', '').strip()
    marca = request.GET.get('marca', '').strip()
    categoria_id = request.GET.get('categoriaId', '').strip()
    estado = request.GET.get('estado', '').strip()
    stock_filter = request.GET.get('stock', '').strip()
    precio_min = request.GET.get('precioMin', '').strip()
    precio_max = request.GET.get('precioMax', '').strip()
    ordenar_por = request.GET.get('ordenarPor', 'nombre').strip()
    direccion = request.GET.get('direccion', 'asc').strip()

    qs = Producto.objects.select_related('categoria').all()
    if producto_id: qs = qs.filter(producto_id=producto_id)
    if nombre: qs = qs.filter(producto_nombre__icontains=nombre)
    if marca: qs = qs.filter(producto_marca__iexact=marca)
    if categoria_id: qs = qs.filter(categoria_id=categoria_id)
    if estado: qs = qs.filter(producto_estado__iexact=estado)
    if precio_min: qs = qs.filter(producto_precio_venta__gte=Decimal(precio_min))
    if precio_max: qs = qs.filter(producto_precio_venta__lte=Decimal(precio_max))

    if stock_filter == 'Agotado':
        ids = Inventario.objects.filter(inventario_stock_actual__lte=0).values_list('producto_id', flat=True)
        qs = qs.filter(producto_id__in=ids)
    elif stock_filter == 'ConStock':
        ids = Inventario.objects.filter(inventario_stock_actual__gt=0).values_list('producto_id', flat=True)
        qs = qs.filter(producto_id__in=ids)

    order_map = {'nombre': 'producto_nombre', 'id': 'producto_id',
                 'precio': 'producto_precio_venta', 'marca': 'producto_marca', 'stock': 'producto_id'}
    of = order_map.get(ordenar_por, 'producto_nombre')
    if direccion == 'desc': of = f'-{of}'
    qs = qs.order_by(of)
    productos = list(qs)

    inv_map = {}
    for inv in Inventario.objects.filter(producto_id__in=[p.producto_id for p in productos]):
        inv_map[inv.producto_id] = inv.inventario_stock_actual or 0
    if ordenar_por == 'stock':
        productos.sort(key=lambda p: inv_map.get(p.producto_id, 0), reverse=(direccion == 'desc'))

    elements = []
    filtros = []
    if producto_id: filtros.append(f'ID: {producto_id}')
    if nombre: filtros.append(f'Nombre: {nombre}')
    if marca: filtros.append(f'Marca: {marca}')
    if categoria_id:
        cn = Categoria.objects.filter(categoria_id=categoria_id).values_list('categoria_nombre', flat=True).first()
        filtros.append(f'Categoría: {cn}')
    if estado: filtros.append(f'Estado: {estado}')
    if stock_filter: filtros.append(f'Stock: {stock_filter}')
    if precio_min: filtros.append(f'Precio mín: ${precio_min}')
    if precio_max: filtros.append(f'Precio máx: ${precio_max}')

    _pdf_header(elements, 'Reporte de Productos', '', f'Total: {len(productos)} productos', filtros)

    s = _get_pdf_styles()
    data = [['ID', 'Código', 'Nombre', 'Marca', 'Categoría', 'Precio', 'Stock', 'Estado']]
    for p in productos:
        data.append([
            str(p.producto_id), p.producto_codigo_bar or '-',
            Paragraph(p.producto_nombre or '-', s['cell']),
            p.producto_marca or '-',
            p.categoria.categoria_nombre if p.categoria else 'Sin categoría',
            f'${p.producto_precio_venta:,.0f}' if p.producto_precio_venta else '$0',
            str(inv_map.get(p.producto_id, 0)), p.producto_estado or '-',
        ])

    t = Table(data, colWidths=[40, 70, 180, 80, 90, 70, 50, 55], repeatRows=1)
    ts = _table_style()
    ts.add('ALIGN', (0, 1), (0, -1), 'CENTER')
    ts.add('ALIGN', (5, 1), (5, -1), 'RIGHT')
    ts.add('ALIGN', (6, 1), (6, -1), 'CENTER')
    ts.add('ALIGN', (7, 1), (7, -1), 'CENTER')
    t.setStyle(ts)
    elements.append(t)

    buf = _build_pdf(elements)
    resp = HttpResponse(buf, content_type='application/pdf')
    resp['Content-Disposition'] = f'inline; filename="reporte_productos_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf"'
    return resp


# ══════════════════════════════════════════════════════════════════
#  REPORTE DE VENTAS
# ══════════════════════════════════════════════════════════════════

@admin_required
def reporte_ventas_view(request):
    metodos_pago = MetodoPago.objects.all().order_by('metodo_pago_nombre')
    
    # Ranking por Vendedores
    ventas_vendedores = Venta.objects.values('user__name', 'user__role').annotate(
        total_ventas=Count('venta_codigo'),
        ingresos=Sum('venta_total')
    ).order_by('-ingresos')

    ranking_vendedores = []
    for v in ventas_vendedores:
        val = v['ingresos'] or 0
        v['ingresos_fmt'] = f"$ {val:,.0f} COP".replace(',', '.')
        ranking_vendedores.append(v)

    # Ranking por Clientes (Top 15)
    ventas_clientes = Venta.objects.values('comprador__name', 'comprador__email').annotate(
        total_compras=Count('venta_codigo'),
        gastado=Sum('venta_total')
    ).order_by('-gastado')[:15]

    ranking_clientes = []
    for c in ventas_clientes:
        val = c['gastado'] or 0
        c['gastado_fmt'] = f"$ {val:,.0f} COP".replace(',', '.')
        ranking_clientes.append(c)

    return render(request, 'admin_panel/reporte_ventas.html', {
        'metodos_pago': metodos_pago,
        'ranking_vendedores': ranking_vendedores,
        'ranking_clientes': ranking_clientes,
    })


@admin_required
def reporte_ventas_pdf(request):
    fecha_desde = request.GET.get('fechaDesde', '').strip()
    fecha_hasta = request.GET.get('fechaHasta', '').strip()
    cliente = request.GET.get('cliente', '').strip()
    metodo_pago_id = request.GET.get('metodoPago', '').strip()
    total_min = request.GET.get('totalMin', '').strip()
    total_max = request.GET.get('totalMax', '').strip()
    ordenar_por = request.GET.get('ordenarPor', 'fecha').strip()
    direccion = request.GET.get('direccion', 'desc').strip()

    qs = Venta.objects.select_related('comprador', 'user', 'metodo_pago').all()
    if fecha_desde: qs = qs.filter(venta_fecha__gte=fecha_desde)
    if fecha_hasta: qs = qs.filter(venta_fecha__lte=fecha_hasta)
    if cliente: qs = qs.filter(Q(comprador__name__icontains=cliente) | Q(user__name__icontains=cliente))
    if metodo_pago_id: qs = qs.filter(metodo_pago_id=metodo_pago_id)
    if total_min: qs = qs.filter(venta_total__gte=Decimal(total_min))
    if total_max: qs = qs.filter(venta_total__lte=Decimal(total_max))

    order_map = {'fecha': 'venta_fecha', 'codigo': 'venta_codigo', 'total': 'venta_total', 'cliente': 'comprador__name'}
    of = order_map.get(ordenar_por, 'venta_fecha')
    if direccion == 'desc': of = f'-{of}'
    qs = qs.order_by(of)
    ventas = list(qs)

    total_general = sum(v.venta_total or 0 for v in ventas)

    elements = []
    filtros = []
    if fecha_desde: filtros.append(f'Desde: {fecha_desde}')
    if fecha_hasta: filtros.append(f'Hasta: {fecha_hasta}')
    if cliente: filtros.append(f'Cliente: {cliente}')
    if metodo_pago_id:
        mp = MetodoPago.objects.filter(metodo_pago_id=metodo_pago_id).values_list('metodo_pago_nombre', flat=True).first()
        filtros.append(f'Método: {mp}')
    if total_min: filtros.append(f'Total mín: ${total_min}')
    if total_max: filtros.append(f'Total máx: ${total_max}')

    _pdf_header(elements, 'Reporte de Ventas', '', f'Total: {len(ventas)} ventas — Ingresos: ${total_general:,.0f}', filtros)

    s = _get_pdf_styles()
    data = [['#', 'Fecha', 'Hora', 'Cliente', 'Vendedor', 'Método Pago', 'Total']]
    for v in ventas:
        data.append([
            str(v.venta_codigo),
            str(v.venta_fecha) if v.venta_fecha else '-',
            v.venta_hora.strftime('%H:%M') if v.venta_hora else '-',
            Paragraph(v.comprador.name if v.comprador else '-', s['cell']),
            v.user.name if v.user else '-',
            v.metodo_pago.metodo_pago_nombre if v.metodo_pago else '-',
            f'${v.venta_total:,.0f}' if v.venta_total else '$0',
        ])

    # Fila de total
    data.append(['', '', '', '', '', 'TOTAL:', f'${total_general:,.0f}'])

    t = Table(data, colWidths=[50, 75, 50, 140, 100, 90, 80], repeatRows=1)
    ts = _table_style()
    ts.add('ALIGN', (0, 1), (0, -1), 'CENTER')
    ts.add('ALIGN', (6, 1), (6, -1), 'RIGHT')
    ts.add('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold')
    ts.add('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e8eaf6'))
    t.setStyle(ts)
    elements.append(t)

    # =============== RESÚMENES ===============
    vendedores_dict = {}
    clientes_dict = {}

    for v in ventas:
        v_name = v.user.name if v.user else 'Desconocido'
        v_role = v.user.role if v.user else 'Desconocido'
        if v_name not in vendedores_dict:
            vendedores_dict[v_name] = {'role': v_role, 'ventas': 0, 'total': Decimal('0')}
        vendedores_dict[v_name]['ventas'] += 1
        vendedores_dict[v_name]['total'] += (v.venta_total or Decimal('0'))

        c_name = v.comprador.name if v.comprador else 'Desconocido'
        if c_name not in clientes_dict:
            clientes_dict[c_name] = {'ventas': 0, 'total': Decimal('0')}
        clientes_dict[c_name]['ventas'] += 1
        clientes_dict[c_name]['total'] += (v.venta_total or Decimal('0'))

    vendedores_list = sorted(vendedores_dict.items(), key=lambda x: x[1]['total'], reverse=True)
    clientes_list = sorted(clientes_dict.items(), key=lambda x: x[1]['total'], reverse=True)[:15]

    elements.append(Spacer(1, 25))
    s_title = ParagraphStyle('T2', parent=s['title'], fontSize=14)
    elements.append(Paragraph('Rendimiento por Vendedor / Origen', s_title))
    
    data_vend = [['Usuario / Canal', 'Rol', 'Ventas', 'Ingresos Generados']]
    for name, info in vendedores_list:
        if info['role'] == 'ADMIN': r_str = 'Admin'
        elif info['role'] == 'VENDEDOR': r_str = 'Vendedor'
        else: r_str = 'Web (Autoservicio)'

        data_vend.append([
            Paragraph(name, s['cell']),
            r_str,
            str(info['ventas']),
            f"$ {info['total']:,.0f} COP".replace(',', '.')
        ])
    t_vend = Table(data_vend, colWidths=[200, 100, 80, 100], repeatRows=1)
    ts_v = _table_style()
    ts_v.add('ALIGN', (2, 1), (2, -1), 'CENTER')
    ts_v.add('ALIGN', (3, 1), (3, -1), 'RIGHT')
    t_vend.setStyle(ts_v)
    elements.append(t_vend)

    elements.append(Spacer(1, 25))
    elements.append(Paragraph('Mejores Clientes (Top Compradores)', s_title))
    
    data_cli = [['Cliente', 'Compras', 'Total Gastado']]
    for name, info in clientes_list:
        data_cli.append([
            Paragraph(name, s['cell']),
            str(info['ventas']),
            f"$ {info['total']:,.0f} COP".replace(',', '.')
        ])
    t_cli = Table(data_cli, colWidths=[300, 80, 100], repeatRows=1)
    ts_c = _table_style()
    ts_c.add('ALIGN', (1, 1), (1, -1), 'CENTER')
    ts_c.add('ALIGN', (2, 1), (2, -1), 'RIGHT')
    t_cli.setStyle(ts_c)
    elements.append(t_cli)
    # =========================================

    buf = _build_pdf(elements)
    resp = HttpResponse(buf, content_type='application/pdf')
    resp['Content-Disposition'] = f'inline; filename="reporte_ventas_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf"'
    return resp


# ══════════════════════════════════════════════════════════════════
#  REPORTE DE INVENTARIO
# ══════════════════════════════════════════════════════════════════

@admin_required
def reporte_inventario_view(request):
    return render(request, 'admin_panel/reporte_inventario.html')


@admin_required
def reporte_inventario_pdf(request):
    stock_filter = request.GET.get('stock', '').strip()
    ubicacion = request.GET.get('ubicacion', '').strip()
    nombre = request.GET.get('nombre', '').strip()
    ordenar_por = request.GET.get('ordenarPor', 'producto').strip()
    direccion = request.GET.get('direccion', 'asc').strip()

    qs = Inventario.objects.select_related('producto', 'producto__categoria').all()
    if stock_filter == 'Agotado': qs = qs.filter(inventario_stock_actual__lte=0)
    elif stock_filter == 'Bajo': qs = qs.filter(inventario_stock_actual__lte=F('inventario_stock_minimo'), inventario_stock_actual__gt=0)
    elif stock_filter == 'Normal': qs = qs.filter(inventario_stock_actual__gt=F('inventario_stock_minimo'))
    if ubicacion: qs = qs.filter(inventario_ubicacion__icontains=ubicacion)
    if nombre: qs = qs.filter(producto__producto_nombre__icontains=nombre)

    order_map = {'producto': 'producto__producto_nombre', 'stock': 'inventario_stock_actual', 'ubicacion': 'inventario_ubicacion'}
    of = order_map.get(ordenar_por, 'producto__producto_nombre')
    if direccion == 'desc': of = f'-{of}'
    qs = qs.order_by(of)
    inventarios = list(qs)

    elements = []
    filtros = []
    if stock_filter: filtros.append(f'Stock: {stock_filter}')
    if ubicacion: filtros.append(f'Ubicación: {ubicacion}')
    if nombre: filtros.append(f'Producto: {nombre}')

    _pdf_header(elements, 'Reporte de Inventario', '', f'Total: {len(inventarios)} registros', filtros)

    s = _get_pdf_styles()
    data = [['ID', 'Producto', 'Categoría', 'Stock Actual', 'Stock Mín', 'Stock Máx', 'Ubicación', 'Estado']]
    for inv in inventarios:
        stock = inv.inventario_stock_actual or 0
        stock_min = inv.inventario_stock_minimo or 0
        if stock <= 0:
            estado_stock = 'AGOTADO'
        elif stock <= stock_min:
            estado_stock = 'BAJO'
        else:
            estado_stock = 'OK'
        data.append([
            str(inv.inventario_id),
            Paragraph(inv.producto.producto_nombre if inv.producto else '-', s['cell']),
            inv.producto.categoria.categoria_nombre if inv.producto and inv.producto.categoria else '-',
            str(stock), str(stock_min), str(inv.inventario_stock_maximo or 0),
            inv.inventario_ubicacion or '-', estado_stock,
        ])

    t = Table(data, colWidths=[40, 160, 80, 60, 60, 60, 80, 55], repeatRows=1)
    ts = _table_style()
    ts.add('ALIGN', (0, 1), (0, -1), 'CENTER')
    ts.add('ALIGN', (3, 1), (5, -1), 'CENTER')
    ts.add('ALIGN', (7, 1), (7, -1), 'CENTER')
    t.setStyle(ts)
    elements.append(t)

    buf = _build_pdf(elements)
    resp = HttpResponse(buf, content_type='application/pdf')
    resp['Content-Disposition'] = f'inline; filename="reporte_inventario_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf"'
    return resp


# ══════════════════════════════════════════════════════════════════
#  REPORTE DE USUARIOS
# ══════════════════════════════════════════════════════════════════

@admin_required
def reporte_usuarios_view(request):
    return render(request, 'admin_panel/reporte_usuarios.html')


@admin_required
def reporte_usuarios_pdf(request):
    role = request.GET.get('role', '').strip()
    enabled = request.GET.get('enabled', '').strip()
    nombre = request.GET.get('nombre', '').strip()
    ordenar_por = request.GET.get('ordenarPor', 'nombre').strip()
    direccion = request.GET.get('direccion', 'asc').strip()

    qs = User.objects.all()
    if role: qs = qs.filter(role__iexact=role)
    if enabled == '1': qs = qs.filter(enabled=True)
    elif enabled == '0': qs = qs.filter(enabled=False)
    if nombre: qs = qs.filter(Q(name__icontains=nombre) | Q(email__icontains=nombre))

    order_map = {'nombre': 'name', 'email': 'email', 'id': 'id', 'rol': 'role'}
    of = order_map.get(ordenar_por, 'name')
    if direccion == 'desc': of = f'-{of}'
    qs = qs.order_by(of)
    usuarios = list(qs)

    elements = []
    filtros = []
    if role: filtros.append(f'Rol: {role}')
    if enabled: filtros.append(f'Estado: {"Activo" if enabled == "1" else "Inactivo"}')
    if nombre: filtros.append(f'Búsqueda: {nombre}')

    _pdf_header(elements, 'Reporte de Usuarios', '', f'Total: {len(usuarios)} usuarios', filtros)

    data = [['ID', 'Nombre', 'Email', 'Rol', 'Estado', 'Fecha Registro']]
    s = _get_pdf_styles()
    for u in usuarios:
        data.append([
            str(u.id),
            Paragraph(u.name or '-', s['cell']),
            Paragraph(u.email or '-', s['cell']),
            u.role or '-',
            'Activo' if u.enabled else 'Inactivo',
            u.created_at.strftime('%d/%m/%Y') if u.created_at else '-',
        ])

    t = Table(data, colWidths=[40, 140, 180, 60, 60, 80], repeatRows=1)
    ts = _table_style()
    ts.add('ALIGN', (0, 1), (0, -1), 'CENTER')
    ts.add('ALIGN', (3, 1), (4, -1), 'CENTER')
    ts.add('ALIGN', (5, 1), (5, -1), 'CENTER')
    t.setStyle(ts)
    elements.append(t)

    buf = _build_pdf(elements)
    resp = HttpResponse(buf, content_type='application/pdf')
    resp['Content-Disposition'] = f'inline; filename="reporte_usuarios_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf"'
    return resp


# ══════════════════════════════════════════════════════════════════
#  REPORTE DE CLIENTES
# ══════════════════════════════════════════════════════════════════

@admin_required
def reporte_clientes_view(request):
    return render(request, 'admin_panel/reporte_clientes.html')


@admin_required
def reporte_clientes_pdf(request):
    nombre = request.GET.get('nombre', '').strip()
    email = request.GET.get('email', '').strip()
    ordenar_por = request.GET.get('ordenarPor', 'nombre').strip()
    direccion = request.GET.get('direccion', 'asc').strip()

    # Consultar usuarios que tienen compras
    compradores_ids = Venta.objects.exclude(comprador__isnull=True).values_list('comprador_id', flat=True).distinct()
    qs = User.objects.filter(id__in=compradores_ids)
    if nombre: qs = qs.filter(name__icontains=nombre)
    if email: qs = qs.filter(email__icontains=email)

    order_map = {'nombre': 'name', 'email': 'email', 'id': 'id'}
    of = order_map.get(ordenar_por, 'name')
    if direccion == 'desc': of = f'-{of}'
    qs = qs.order_by(of)
    clientes = list(qs)

    # Contar compras por comprador
    compras_map = {}
    for v in Venta.objects.values('comprador_id').annotate(total=Count('venta_codigo'), suma=Sum('venta_total')):
        compras_map[v['comprador_id']] = {'total': v['total'], 'suma': v['suma'] or 0}

    elements = []
    filtros = []
    if nombre: filtros.append(f'Nombre: {nombre}')
    if email: filtros.append(f'Email: {email}')

    _pdf_header(elements, 'Reporte de Clientes', '', f'Total: {len(clientes)} clientes', filtros)

    data = [['ID', 'Nombre', 'Documento', 'Email', 'Teléfono', 'Dirección', 'Compras', 'Total Compras']]
    s = _get_pdf_styles()
    for c in clientes:
        info = compras_map.get(c.id, {'total': 0, 'suma': 0})
        data.append([
            str(c.id),
            Paragraph(c.name or '-', s['cell']),
            c.numero_documento or '-',
            Paragraph(c.email or '-', s['cell']),
            c.telefono or '-',
            Paragraph(c.direccion or '-', s['cell']),
            str(info['total']),
            f'${info["suma"]:,.0f}',
        ])

    t = Table(data, colWidths=[35, 110, 70, 110, 65, 100, 45, 70], repeatRows=1)
    ts = _table_style()
    ts.add('ALIGN', (0, 1), (0, -1), 'CENTER')
    ts.add('ALIGN', (6, 1), (6, -1), 'CENTER')
    ts.add('ALIGN', (7, 1), (7, -1), 'RIGHT')
    t.setStyle(ts)
    elements.append(t)

    buf = _build_pdf(elements)
    resp = HttpResponse(buf, content_type='application/pdf')
    resp['Content-Disposition'] = f'inline; filename="reporte_clientes_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf"'
    return resp


# ══════════════════════════════════════════════════════════════════
#  REPORTE DE VENDEDORES
# ══════════════════════════════════════════════════════════════════

@admin_required
def reporte_vendedores_view(request):
    """Vista HTML del reporte de vendedores con sus ventas."""
    vendedores = User.objects.filter(role__in=['ADMIN', 'VENDEDOR'], enabled=True).order_by('name')

    vendedores_data = []
    for v in vendedores:
        ventas_qs = Venta.objects.filter(user=v)
        total_ventas = ventas_qs.count()
        ingresos = ventas_qs.aggregate(total=Sum('venta_total'))['total'] or 0
        vendedores_data.append({
            'id': v.id,
            'name': v.name,
            'email': v.email,
            'role': v.role,
            'total_ventas': total_ventas,
            'ingresos': ingresos,
            'ingresos_fmt': f"${ingresos:,.0f}",
        })

    vendedores_data.sort(key=lambda x: x['ingresos'], reverse=True)

    return render(request, 'admin_panel/reporte_vendedores.html', {
        'vendedores_data': vendedores_data,
        'vendedores': vendedores,
    })


@admin_required
def reporte_vendedores_pdf(request):
    """Genera PDF con rendimiento de vendedores."""
    vendedor_id = request.GET.get('vendedor', '').strip()
    fecha_desde = request.GET.get('fechaDesde', '').strip()
    fecha_hasta = request.GET.get('fechaHasta', '').strip()

    # Obtener vendedores (ADMIN y VENDEDOR)
    vendedores_qs = User.objects.filter(role__in=['ADMIN', 'VENDEDOR'], enabled=True)
    if vendedor_id:
        vendedores_qs = vendedores_qs.filter(id=vendedor_id)

    vendedores_data = []
    total_ventas_general = 0
    total_ingresos_general = Decimal('0')

    for v in vendedores_qs.order_by('name'):
        ventas_qs = Venta.objects.filter(user=v)
        if fecha_desde:
            ventas_qs = ventas_qs.filter(venta_fecha__gte=fecha_desde)
        if fecha_hasta:
            ventas_qs = ventas_qs.filter(venta_fecha__lte=fecha_hasta)

        total_ventas = ventas_qs.count()
        ingresos = ventas_qs.aggregate(total=Sum('venta_total'))['total'] or Decimal('0')

        vendedores_data.append({
            'user': v,
            'total_ventas': total_ventas,
            'ingresos': ingresos,
        })
        total_ventas_general += total_ventas
        total_ingresos_general += ingresos

    vendedores_data.sort(key=lambda x: x['ingresos'], reverse=True)

    elements = []
    filtros = []
    if vendedor_id:
        vn = User.objects.filter(id=vendedor_id).values_list('name', flat=True).first()
        filtros.append(f'Vendedor: {vn}')
    if fecha_desde:
        filtros.append(f'Desde: {fecha_desde}')
    if fecha_hasta:
        filtros.append(f'Hasta: {fecha_hasta}')

    _pdf_header(
        elements, 'Reporte de Vendedores', '',
        f'Total: {len(vendedores_data)} vendedores — {total_ventas_general} ventas — Ingresos: ${total_ingresos_general:,.0f}',
        filtros
    )

    s = _get_pdf_styles()
    data = [['#', 'Vendedor', 'Email', 'Rol', 'Total Ventas', 'Ingresos Generados']]

    for i, vd in enumerate(vendedores_data, 1):
        v = vd['user']
        role_str = 'Admin' if v.role == 'ADMIN' else 'Vendedor'
        data.append([
            str(i),
            Paragraph(v.name or '-', s['cell']),
            Paragraph(v.email or '-', s['cell']),
            role_str,
            str(vd['total_ventas']),
            f"${vd['ingresos']:,.0f}",
        ])

    # Fila de total
    data.append(['', '', '', 'TOTAL:', str(total_ventas_general), f'${total_ingresos_general:,.0f}'])

    t = Table(data, colWidths=[30, 140, 160, 70, 70, 100], repeatRows=1)
    ts = _table_style()
    ts.add('ALIGN', (0, 1), (0, -1), 'CENTER')
    ts.add('ALIGN', (3, 1), (3, -1), 'CENTER')
    ts.add('ALIGN', (4, 1), (4, -1), 'CENTER')
    ts.add('ALIGN', (5, 1), (5, -1), 'RIGHT')
    ts.add('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold')
    ts.add('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e8eaf6'))
    t.setStyle(ts)
    elements.append(t)

    # Detalle por vendedor si se filtra uno específico
    if vendedor_id and vendedores_data:
        vd = vendedores_data[0]
        ventas_detalle = Venta.objects.filter(user=vd['user']).select_related('comprador', 'metodo_pago')
        if fecha_desde:
            ventas_detalle = ventas_detalle.filter(venta_fecha__gte=fecha_desde)
        if fecha_hasta:
            ventas_detalle = ventas_detalle.filter(venta_fecha__lte=fecha_hasta)
        ventas_detalle = ventas_detalle.order_by('-venta_fecha', '-venta_hora')

        if ventas_detalle.exists():
            elements.append(Spacer(1, 25))
            s_title = ParagraphStyle('T2', parent=s['title'], fontSize=14)
            elements.append(Paragraph(f'Detalle de Ventas — {vd["user"].name}', s_title))

            data_det = [['#', 'Fecha', 'Hora', 'Cliente', 'Método Pago', 'Total']]
            for venta in ventas_detalle:
                data_det.append([
                    str(venta.venta_codigo),
                    str(venta.venta_fecha) if venta.venta_fecha else '-',
                    venta.venta_hora.strftime('%H:%M') if venta.venta_hora else '-',
                    Paragraph(venta.comprador.name if venta.comprador else '-', s['cell']),
                    venta.metodo_pago.metodo_pago_nombre if venta.metodo_pago else '-',
                    f'${venta.venta_total:,.0f}' if venta.venta_total else '$0',
                ])
            t_det = Table(data_det, colWidths=[50, 75, 50, 160, 100, 80], repeatRows=1)
            ts_d = _table_style()
            ts_d.add('ALIGN', (0, 1), (0, -1), 'CENTER')
            ts_d.add('ALIGN', (5, 1), (5, -1), 'RIGHT')
            t_det.setStyle(ts_d)
            elements.append(t_det)

    buf = _build_pdf(elements)
    resp = HttpResponse(buf, content_type='application/pdf')
    resp['Content-Disposition'] = f'inline; filename="reporte_vendedores_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf"'
    return resp


# ══════════════════════════════════════════════════════════════════
#  CARGA MASIVA DE PRODUCTOS
# ══════════════════════════════════════════════════════════════════

@admin_required
def carga_masiva_view(request):
    categorias = Categoria.objects.filter(estado=1).order_by('categoria_nombre')
    if request.method == 'POST':
        archivo = request.FILES.get('archivo')
        if not archivo:
            messages.error(request, 'Debes seleccionar un archivo.')
            return render(request, 'admin_panel/carga_masiva.html', {'categorias': categorias, 'tipo': 'productos'})

        filename = archivo.name.lower()
        if not (filename.endswith('.csv') or filename.endswith('.xlsx') or filename.endswith('.xls')):
            messages.error(request, 'Formato no soportado. Usa archivos .csv o .xlsx')
            return render(request, 'admin_panel/carga_masiva.html', {'categorias': categorias, 'tipo': 'productos'})

        try:
            rows = _parse_csv(archivo) if filename.endswith('.csv') else _parse_excel(archivo)
            creados, errores = _procesar_carga_productos(rows)
            if errores:
                for err in errores:
                    messages.error(request, err)
                messages.warning(request, f'Se encontraron {len(errores)} error(es). No se importó ningún producto. Corrige los errores y vuelve a intentarlo.')
                return render(request, 'admin_panel/carga_masiva.html', {'categorias': categorias, 'tipo': 'productos', 'hay_errores': True})
            if creados > 0:
                messages.success(request, f'¡{creados} producto(s) importado(s) exitosamente!')
        except Exception as e:
            messages.error(request, f'Error al procesar el archivo: {str(e)}')
            return render(request, 'admin_panel/carga_masiva.html', {'categorias': categorias, 'tipo': 'productos', 'hay_errores': True})

        return render(request, 'admin_panel/carga_masiva.html', {'categorias': categorias, 'carga_completada': True, 'tipo': 'productos'})

    return render(request, 'admin_panel/carga_masiva.html', {'categorias': categorias, 'tipo': 'productos'})


@admin_required
def descargar_plantilla_csv(request):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="plantilla_productos.csv"'
    response.write('\ufeff')
    writer = csv.writer(response, delimiter=';')
    writer.writerow(['nombre', 'codigo', 'marca', 'precio', 'unidad', 'categoria', 'estado', 'stock', 'stock_minimo', 'stock_maximo', 'ubicacion', 'resumen'])
    writer.writerow(['Laptop HP Pavilion', 'HP-PAV-001', 'HP', '2500000', 'Unidad', 'Computadores', 'Activo', '15', '5', '50', 'Bodega A-1', 'Laptop HP Pavilion 15'])
    writer.writerow(['Mouse Logitech G502', 'LOG-G502', 'Logitech', '180000', 'Unidad', 'Accesorios', 'Activo', '30', '10', '100', 'Bodega B-2', 'Mouse gaming Logitech'])
    return response


@admin_required
def descargar_plantilla_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Productos'
    hf = Font(bold=True, color='FFFFFF', size=11)
    hfill = PatternFill(start_color='0D47A1', end_color='0D47A1', fill_type='solid')
    ha = Alignment(horizontal='center', vertical='center', wrap_text=True)
    tb = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ['nombre', 'codigo', 'marca', 'precio', 'unidad', 'categoria', 'estado', 'stock', 'stock_minimo', 'stock_maximo', 'ubicacion', 'resumen']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font, cell.fill, cell.alignment, cell.border = hf, hfill, ha, tb

    examples = [
        ['Laptop HP Pavilion', 'HP-PAV-001', 'HP', 2500000, 'Unidad', 'Computadores', 'Activo', 15, 5, 50, 'Bodega A-1', 'Laptop HP Pavilion 15'],
        ['Mouse Logitech G502', 'LOG-G502', 'Logitech', 180000, 'Unidad', 'Accesorios', 'Activo', 30, 10, 100, 'Bodega B-2', 'Mouse gaming Logitech'],
    ]
    for ri, row in enumerate(examples, 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = tb

    widths = [25, 15, 15, 12, 10, 18, 10, 8, 12, 12, 15, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="plantilla_productos.xlsx"'
    return resp


# ══════════════════════════════════════════════════════════════════
#  CARGA MASIVA DE USUARIOS
# ══════════════════════════════════════════════════════════════════

@admin_required
def carga_masiva_usuarios_view(request):
    if request.method == 'POST':
        archivo = request.FILES.get('archivo')
        if not archivo:
            messages.error(request, 'Debes seleccionar un archivo.')
            return render(request, 'admin_panel/carga_masiva_usuarios.html')

        filename = archivo.name.lower()
        if not (filename.endswith('.csv') or filename.endswith('.xlsx') or filename.endswith('.xls')):
            messages.error(request, 'Formato no soportado. Usa archivos .csv o .xlsx')
            return render(request, 'admin_panel/carga_masiva_usuarios.html')

        try:
            rows = _parse_csv(archivo) if filename.endswith('.csv') else _parse_excel(archivo)
            creados, errores = _procesar_carga_usuarios(rows)
            if errores:
                for err in errores:
                    messages.error(request, err)
                messages.warning(request, f'Se encontraron {len(errores)} error(es). No se importó ningún usuario. Corrige los errores y vuelve a intentarlo.')
                return render(request, 'admin_panel/carga_masiva_usuarios.html', {'hay_errores': True})
            if creados > 0:
                messages.success(request, f'¡{creados} usuario(s) importado(s) exitosamente!')
        except Exception as e:
            messages.error(request, f'Error al procesar el archivo: {str(e)}')
            return render(request, 'admin_panel/carga_masiva_usuarios.html', {'hay_errores': True})

        return render(request, 'admin_panel/carga_masiva_usuarios.html', {'carga_completada': True})

    return render(request, 'admin_panel/carga_masiva_usuarios.html')


@admin_required
def descargar_plantilla_usuarios_csv(request):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="plantilla_usuarios.csv"'
    response.write('\ufeff')
    writer = csv.writer(response, delimiter=';')
    writer.writerow(['nombre', 'email', 'password', 'rol'])
    writer.writerow(['Juan Pérez', 'juan@email.com', 'MiPass123', 'USER'])
    writer.writerow(['Ana Admin', 'ana@email.com', 'Admin456', 'ADMIN'])
    return response


@admin_required
def descargar_plantilla_usuarios_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Usuarios'
    hf = Font(bold=True, color='FFFFFF', size=11)
    hfill = PatternFill(start_color='0D47A1', end_color='0D47A1', fill_type='solid')
    ha = Alignment(horizontal='center', vertical='center', wrap_text=True)
    tb = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ['nombre', 'email', 'password', 'rol']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font, cell.fill, cell.alignment, cell.border = hf, hfill, ha, tb

    examples = [
        ['Juan Pérez', 'juan@email.com', 'MiPass123', 'USER'],
        ['Ana Admin', 'ana@email.com', 'Admin456', 'ADMIN'],
    ]
    for ri, row in enumerate(examples, 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = tb

    widths = [25, 30, 20, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Instrucciones
    ws2 = wb.create_sheet('Instrucciones')
    instr = [
        ['Campo', 'Descripción', 'Obligatorio'],
        ['nombre', 'Nombre completo del usuario', 'Sí'],
        ['email', 'Correo electrónico (debe ser único)', 'Sí'],
        ['password', 'Contraseña (mínimo 6 caracteres)', 'Sí'],
        ['rol', 'USER o ADMIN (default: USER)', 'No'],
    ]
    for ri, row in enumerate(instr, 1):
        for ci, val in enumerate(row, 1):
            cell = ws2.cell(row=ri, column=ci, value=val)
            if ri == 1:
                cell.font, cell.fill, cell.alignment = hf, hfill, ha
            cell.border = tb
    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['B'].width = 40
    ws2.column_dimensions['C'].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="plantilla_usuarios.xlsx"'
    return resp


# ══════════════════════════════════════════════════════════════════
#  HELPERS INTERNOS
# ══════════════════════════════════════════════════════════════════

def _parse_csv(archivo):
    content = archivo.read().decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(content), delimiter=';')
    fieldnames = reader.fieldnames or []
    if len(fieldnames) <= 1:
        reader = csv.DictReader(io.StringIO(content), delimiter=',')
    return list(reader)


def _parse_excel(archivo):
    import openpyxl
    wb = openpyxl.load_workbook(archivo, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = [str(h).strip().lower() if h else '' for h in next(rows_iter)]
    except StopIteration:
        return []
    rows = []
    for row in rows_iter:
        if all(v is None for v in row):
            continue
        row_dict = {}
        for i, header in enumerate(headers):
            if header and i < len(row):
                val = row[i]
                row_dict[header] = str(val).strip() if val is not None else ''
        rows.append(row_dict)
    wb.close()
    return rows


def _procesar_carga_productos(rows):
    errores = []
    cat_cache = {}
    for cat in Categoria.objects.filter(estado=1):
        cat_cache[cat.categoria_nombre.lower().strip()] = cat

    # Fase 1: Validar todas las filas
    productos_validos = []
    for idx, row in enumerate(rows, start=2):
        nombre = row.get('nombre', '').strip()
        if not nombre:
            errores.append(f'Fila {idx}: El nombre es obligatorio.')
            continue
        precio_str = row.get('precio', '0').strip().replace(',', '').replace('$', '')
        try:
            precio = Decimal(precio_str) if precio_str else Decimal('0')
        except (InvalidOperation, ValueError):
            errores.append(f'Fila {idx}: Precio inválido para "{nombre}".')
            continue

        cat_nombre = row.get('categoria', '').strip()
        categoria = cat_cache.get(cat_nombre.lower()) if cat_nombre else None
        if cat_nombre and not categoria:
            errores.append(f'Fila {idx}: Categoría "{cat_nombre}" no encontrada.')
            continue

        estado = row.get('estado', 'Activo').strip()
        if estado not in ('Activo', 'Inactivo'): estado = 'Activo'

        productos_validos.append({
            'nombre': nombre, 'codigo': row.get('codigo', '').strip(),
            'marca': row.get('marca', '').strip(), 'precio': precio,
            'unidad': row.get('unidad', 'Unidad').strip() or 'Unidad',
            'categoria': categoria, 'estado': estado, 'resumen': row.get('resumen', '').strip(),
            'stock': _safe_int(row.get('stock', '0')),
            'stock_minimo': _safe_int(row.get('stock_minimo', '0')),
            'stock_maximo': _safe_int(row.get('stock_maximo', '0')),
            'ubicacion': row.get('ubicacion', '').strip(),
        })

    # Si hay errores, no crear nada
    if errores:
        return 0, errores

    # Fase 2: Crear todos los registros
    creados = 0
    for p in productos_validos:
        try:
            producto = Producto.objects.create(
                producto_nombre=p['nombre'], producto_codigo_bar=p['codigo'],
                producto_marca=p['marca'], producto_precio_venta=p['precio'],
                tipo_unidad=p['unidad'], categoria=p['categoria'],
                producto_estado=p['estado'], resumen=p['resumen'],
            )
            Inventario.objects.create(
                producto=producto, inventario_stock_actual=p['stock'],
                inventario_stock_minimo=p['stock_minimo'],
                inventario_stock_maximo=p['stock_maximo'],
                inventario_ubicacion=p['ubicacion'],
            )
            creados += 1
        except Exception as e:
            errores.append(f'Error creando "{p["nombre"]}": {str(e)}')

    return creados, errores


def _procesar_carga_usuarios(rows):
    errores = []
    emails_existentes = set(e.lower() for e in User.objects.values_list('email', flat=True))
    emails_en_archivo = set()

    # Fase 1: Validar todas las filas
    usuarios_validos = []
    for idx, row in enumerate(rows, start=2):
        nombre = row.get('nombre', '').strip()
        email = row.get('email', '').strip()
        password = row.get('password', '').strip()

        if not nombre:
            errores.append(f'Fila {idx}: El nombre es obligatorio.')
            continue
        if not email:
            errores.append(f'Fila {idx}: El email es obligatorio para "{nombre}".')
            continue
        if not password or len(password) < 6:
            errores.append(f'Fila {idx}: La contraseña debe tener al menos 6 caracteres para "{nombre}".')
            continue
        if email.lower() in emails_existentes:
            errores.append(f'Fila {idx}: El email "{email}" ya está registrado en el sistema.')
            continue
        if email.lower() in emails_en_archivo:
            errores.append(f'Fila {idx}: El email "{email}" está duplicado en el archivo.')
            continue

        role = row.get('rol', 'USER').strip().upper()
        if role not in ('USER', 'ADMIN'):
            errores.append(f'Fila {idx}: Rol "{row.get("rol", "").strip()}" no válido para "{nombre}". Usa USER o ADMIN.')
            continue

        emails_en_archivo.add(email.lower())
        usuarios_validos.append({
            'nombre': nombre, 'email': email,
            'password': password, 'role': role,
        })

    # Si hay errores, no crear nada
    if errores:
        return 0, errores

    # Fase 2: Crear todos los registros
    creados = 0
    for u in usuarios_validos:
        try:
            hashed = bcrypt.hashpw(u['password'].encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
            User.objects.create(
                name=u['nombre'], email=u['email'], password=hashed,
                role=u['role'], enabled=True,
                created_at=datetime.now(), updated_at=datetime.now(),
            )
            creados += 1
        except Exception as e:
            errores.append(f'Error creando "{u["nombre"]}": {str(e)}')

    return creados, errores


def _safe_int(value):
    try:
        return int(float(str(value).strip().replace(',', ''))) if value else 0
    except (ValueError, TypeError):
        return 0
